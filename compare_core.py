# compare_core.py
import os
import re
from typing import Optional, Union, Tuple, List

import pandas as pd

try:
    import openpyxl
except Exception:
    openpyxl = None


def is_excel(path: str) -> bool:
    return os.path.splitext(path.lower())[1] in [".xlsx", ".xlsm", ".xls"]


def is_csv(path: str) -> bool:
    return os.path.splitext(path.lower())[1] == ".csv"


def index_to_excel_col_letter(idx0: int) -> str:
    n = idx0 + 1
    letters = ""
    while n:
        n, rem = divmod(n - 1, 26)
        letters = chr(rem + ord("A")) + letters
    return letters


def excel_col_letter_to_index(col: str) -> int:
    col = col.strip().upper()
    if not re.fullmatch(r"[A-Z]+", col):
        raise ValueError(f"Invalid Excel column letter: {col}")
    idx = 0
    for ch in col:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def parse_sheet_spec(sheet: Optional[str]) -> Union[str, int, None]:
    if sheet is None or str(sheet).strip() == "":
        return None
    sh = str(sheet).strip()
    if re.fullmatch(r"\d+", sh):
        idx = int(sh)
        if idx <= 0:
            raise ValueError("Sheet number must be >= 1")
        return idx - 1
    return sh


def get_excel_sheet_names(path: str) -> List[str]:
    if openpyxl is None:
        return pd.ExcelFile(path).sheet_names
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def get_excel_last_col_letter(path: str, sheet: Union[str, int, None]) -> Optional[str]:
    if openpyxl is None:
        return None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is None:
            ws = wb.worksheets[0]
        elif isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]
        max_col = ws.max_column or 1
        return index_to_excel_col_letter(max_col - 1)
    finally:
        wb.close()


def read_excel_auto_usecols(path: str, sheet: Union[str, int, None], header: Optional[int]) -> pd.DataFrame:
    last = get_excel_last_col_letter(path, sheet)
    usecols = f"A:{last}" if last else None
    return pd.read_excel(path, sheet_name=sheet, header=header, usecols=usecols, engine="openpyxl", dtype="string")


def read_csv(path: str, header: Optional[int]) -> pd.DataFrame:
    try:
        return pd.read_csv(path, header=header, encoding="utf-8-sig", dtype="string")
    except UnicodeDecodeError:
        return pd.read_csv(path, header=header, encoding="cp1256", dtype="string")


def looks_like_bad_header(cols: List[object]) -> bool:
    if not cols:
        return True
    bad = 0
    for c in cols:
        s = str(c).strip()
        if s == "" or s.lower().startswith("unnamed"):
            bad += 1
        elif re.fullmatch(r"\d+(\.\d+)?", s):
            bad += 1
    return (bad / max(1, len(cols))) >= 0.5


def auto_detect_header_and_load(path: str, sheet: Union[str, int, None]) -> Tuple[pd.DataFrame, Optional[int], str]:
    if is_csv(path):
        df0 = read_csv(path, header=0)
        dfN = read_csv(path, header=None)
    else:
        df0 = read_excel_auto_usecols(path, sheet, header=0)
        dfN = read_excel_auto_usecols(path, sheet, header=None)

    score0 = (0 if looks_like_bad_header(list(df0.columns)) else 2) + min(df0.shape[1], 20) / 20
    scoreN = (2 if looks_like_bad_header(list(df0.columns)) else 0) + min(dfN.shape[1], 20) / 20

    if scoreN > score0:
        return dfN, None, "no_header"
    return df0, 0, "header_0"


def normalize_values(series: pd.Series, case_insensitive: bool, drop_blanks: bool) -> pd.Series:
    s = series.astype("string").copy().str.strip()
    if case_insensitive:
        s = s.str.lower()
    if drop_blanks:
        s = s.replace("", pd.NA)
    return s


def auto_pick_best_column_index(df: pd.DataFrame) -> int:
    best_idx = 0
    best_score = -1.0
    nrows = len(df)
    if nrows == 0:
        return 0

    for i in range(df.shape[1]):
        col = df.iloc[:, i].astype("string").str.strip()
        non_null = col.dropna()
        nn = len(non_null)
        if nn == 0:
            continue
        non_null_ratio = nn / nrows
        uniq = non_null.nunique(dropna=True)
        unique_ratio = uniq / max(1, nn)

        if non_null_ratio < 0.2 and nn < 10:
            continue

        score = (non_null_ratio * 0.6) + (unique_ratio * 0.4)
        if score > best_score:
            best_score = score
            best_idx = i

    return best_idx


def pick_series_by_index_or_name(df: pd.DataFrame, spec: str) -> pd.Series:
    spec = str(spec).strip()
    if re.fullmatch(r"\d+", spec):
        idx = int(spec) - 1
        return df.iloc[:, idx]
    if re.fullmatch(r"[A-Za-z]+", spec):
        idx = excel_col_letter_to_index(spec)
        return df.iloc[:, idx]
    if spec in df.columns:
        return df[spec]
    raise KeyError(f"Column not found: {spec}")


def compare_files(
    file_a: str,
    file_b: str,
    sheet_a: Optional[str] = None,
    sheet_b: Optional[str] = None,
    col_a: Optional[str] = None,
    col_b: Optional[str] = None,
    out_path: str = "compare_result.xlsx",
    case_insensitive: bool = False,
    keep_duplicates: bool = False,
    keep_blanks: bool = False,
) -> dict:
    sa = parse_sheet_spec(sheet_a)
    sb = parse_sheet_spec(sheet_b)

    df_a, header_a, header_mode_a = auto_detect_header_and_load(file_a, sa)
    df_b, header_b, header_mode_b = auto_detect_header_and_load(file_b, sb)

    if not col_a:
        idx = auto_pick_best_column_index(df_a)
        col_a = index_to_excel_col_letter(idx)
    if not col_b:
        idx = auto_pick_best_column_index(df_b)
        col_b = index_to_excel_col_letter(idx)

    ser_a = pick_series_by_index_or_name(df_a, col_a)
    ser_b = pick_series_by_index_or_name(df_b, col_b)

    norm_a = normalize_values(ser_a, case_insensitive, drop_blanks=not keep_blanks)
    norm_b = normalize_values(ser_b, case_insensitive, drop_blanks=not keep_blanks)

    if keep_duplicates:
        list_a = norm_a.dropna().tolist()
        list_b = norm_b.dropna().tolist()
        set_a = set(list_a)
        set_b = set(list_b)
        matched = sorted(set_a & set_b)
        only_a = sorted(set_a - set_b)
        only_b = sorted(set_b - set_a)
        occ_a = pd.DataFrame({"key": list_a})
        occ_b = pd.DataFrame({"key": list_b})
    else:
        set_a = set(norm_a.dropna().unique().tolist())
        set_b = set(norm_b.dropna().unique().tolist())
        matched = sorted(set_a & set_b)
        only_a = sorted(set_a - set_b)
        only_b = sorted(set_b - set_a)
        occ_a = occ_b = None

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame({"key": matched}).to_excel(writer, index=False, sheet_name="Matched")
        pd.DataFrame({"key": only_a}).to_excel(writer, index=False, sheet_name="OnlyInA")
        pd.DataFrame({"key": only_b}).to_excel(writer, index=False, sheet_name="OnlyInB")

        meta = pd.DataFrame(
            {
                "item": [
                    "file_a","file_b","sheet_a","sheet_b",
                    "header_a_auto","header_b_auto",
                    "col_a_used","col_b_used",
                    "case_insensitive","unique_only","blanks_dropped",
                    "count_matched","count_only_in_a","count_only_in_b",
                ],
                "value": [
                    file_a,file_b,str(sheet_a),str(sheet_b),
                    str(header_a),str(header_b),
                    str(col_a),str(col_b),
                    str(case_insensitive),str(not keep_duplicates),str(not keep_blanks),
                    str(len(matched)),str(len(only_a)),str(len(only_b)),
                ],
            }
        )
        meta.to_excel(writer, index=False, sheet_name="Meta")

        if occ_a is not None and occ_b is not None:
            occ_a.to_excel(writer, index=False, sheet_name="A_Occurrences")
            occ_b.to_excel(writer, index=False, sheet_name="B_Occurrences")

    return {
        "out": out_path,
        "mode": "compare",
        "matched": len(matched),
        "only_a": len(only_a),
        "only_b": len(only_b),
        "col_a": col_a,
        "col_b": col_b,
        "header_mode_a": header_mode_a,
        "header_mode_b": header_mode_b,
        "df_a_shape": df_a.shape,
        "df_b_shape": df_b.shape,
    }


def _dedupe_b(df_b: pd.DataFrame, key_col_name: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    tmp = df_b.copy()
    dup_mask = tmp[key_col_name].duplicated(keep=False) & tmp[key_col_name].notna()
    dup_report = tmp.loc[dup_mask].copy()
    first = tmp.dropna(subset=[key_col_name]).drop_duplicates(subset=[key_col_name], keep="first").copy()
    return first, dup_report


def xlookup_join(
    file_a: str,
    file_b: str,
    sheet_a: Optional[str] = None,
    sheet_b: Optional[str] = None,
    col_a: Optional[str] = None,
    col_b: Optional[str] = None,
    b_return_cols: Optional[list[str]] = None,
    out_path: str = "lookup_result.xlsx",
    case_insensitive: bool = False,
    keep_blanks: bool = False,
) -> dict:
    sa = parse_sheet_spec(sheet_a)
    sb = parse_sheet_spec(sheet_b)

    df_a, header_a, header_mode_a = auto_detect_header_and_load(file_a, sa)
    df_b, header_b, header_mode_b = auto_detect_header_and_load(file_b, sb)

    if not col_a:
        idx = auto_pick_best_column_index(df_a)
        col_a = index_to_excel_col_letter(idx)
    if not col_b:
        idx = auto_pick_best_column_index(df_b)
        col_b = index_to_excel_col_letter(idx)

    ser_a = pick_series_by_index_or_name(df_a, col_a)
    ser_b = pick_series_by_index_or_name(df_b, col_b)

    key_a = normalize_values(ser_a, case_insensitive, drop_blanks=not keep_blanks)
    key_b = normalize_values(ser_b, case_insensitive, drop_blanks=not keep_blanks)

    a2 = df_a.copy()
    b2 = df_b.copy()
    a2["_key__"] = key_a
    b2["_key__"] = key_b

    if b_return_cols:
        selected = [c for c in b_return_cols if c in b2.columns]
    else:
        selected = [c for c in b2.columns if c != "_key__"]

    b_first, dup_report = _dedupe_b(b2, "_key__")

    lookup_b = b_first[["_key__"] + selected].copy()
    rename_map = {}
    selected_out = []
    for c in selected:
        new_name = f"B__{c}"
        rename_map[c] = new_name
        selected_out.append(new_name)
    lookup_b = lookup_b.rename(columns=rename_map)

    merged = a2.merge(lookup_b, on="_key__", how="left")

    if selected_out:
        not_found = merged[merged[selected_out].isna().all(axis=1) & merged["_key__"].notna()].copy()
    else:
        not_found = merged[merged["_key__"].notna()].copy()

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        merged.drop(columns=["_key__"]).to_excel(writer, index=False, sheet_name="A_with_lookups")
        not_found.drop(columns=["_key__"]).to_excel(writer, index=False, sheet_name="NotFound_in_B")
        if "_key__" in dup_report.columns:
            dup_report = dup_report.drop(columns=["_key__"])
        dup_report.to_excel(writer, index=False, sheet_name="Duplicates_in_B")

        meta = pd.DataFrame(
            {
                "item": [
                    "file_a","file_b",
                    "sheet_a","sheet_b",
                    "header_a_auto","header_b_auto",
                    "col_a_used","col_b_used",
                    "selected_b_cols",
                    "case_insensitive","blanks_dropped",
                    "count_a_rows",
                    "count_not_found",
                    "count_dup_rows_in_b",
                ],
                "value": [
                    file_a,file_b,
                    str(sheet_a),str(sheet_b),
                    str(header_a),str(header_b),
                    str(col_a),str(col_b),
                    ", ".join(map(str, selected_out)),
                    str(bool(case_insensitive)), str(not keep_blanks),
                    str(len(df_a)),
                    str(len(not_found)),
                    str(len(dup_report)),
                ],
            }
        )
        meta.to_excel(writer, index=False, sheet_name="Meta")

    return {
        "out": out_path,
        "mode": "lookup",
        "col_a": col_a,
        "col_b": col_b,
        "selected_b_cols": selected_out,
        "not_found": len(not_found),
        "dup_rows_in_b": len(dup_report),
        "a_rows": len(df_a),
        "header_mode_a": header_mode_a,
        "header_mode_b": header_mode_b,
    }


def differences_report(
    file_a: str,
    file_b: str,
    sheet_a: Optional[str] = None,
    sheet_b: Optional[str] = None,
    col_a: Optional[str] = None,
    col_b: Optional[str] = None,
    compare_cols: Optional[list[str]] = None,
    out_path: str = "diff_result.xlsx",
    case_insensitive: bool = False,
    keep_blanks: bool = False,
) -> dict:
    sa = parse_sheet_spec(sheet_a)
    sb = parse_sheet_spec(sheet_b)

    df_a, header_a, header_mode_a = auto_detect_header_and_load(file_a, sa)
    df_b, header_b, header_mode_b = auto_detect_header_and_load(file_b, sb)

    if not col_a:
        idx = auto_pick_best_column_index(df_a)
        col_a = index_to_excel_col_letter(idx)
    if not col_b:
        idx = auto_pick_best_column_index(df_b)
        col_b = index_to_excel_col_letter(idx)

    key_a = normalize_values(pick_series_by_index_or_name(df_a, col_a), case_insensitive, drop_blanks=not keep_blanks)
    key_b = normalize_values(pick_series_by_index_or_name(df_b, col_b), case_insensitive, drop_blanks=not keep_blanks)

    a2 = df_a.copy()
    b2 = df_b.copy()
    a2["_key__"] = key_a
    b2["_key__"] = key_b

    common = [c for c in df_a.columns if c in df_b.columns]
    cols = [c for c in (compare_cols or common) if c in common]

    # remove key col names if found
    def _colname(df, spec):
        spec = str(spec)
        if re.fullmatch(r"[A-Za-z]+", spec):
            return df.columns[excel_col_letter_to_index(spec)]
        if re.fullmatch(r"\d+", spec):
            return df.columns[int(spec) - 1]
        return spec if spec in df.columns else None
    ka = _colname(df_a, col_a)
    kb = _colname(df_b, col_b)
    if ka in cols: cols.remove(ka)
    if kb in cols: cols.remove(kb)

    b_first, dup_report = _dedupe_b(b2, "_key__")
    merged = a2.merge(b_first, on="_key__", how="left", suffixes=("_A", "_B"))

    b_cols = [c for c in merged.columns if str(c).endswith("_B")]
    if b_cols:
        not_found = merged[merged["_key__"].notna() & merged[b_cols].isna().all(axis=1)].copy()
    else:
        b_keys = set(b_first["_key__"].dropna().tolist())
        not_found = merged[merged["_key__"].notna() & ~merged["_key__"].isin(b_keys)].copy()

    diff_flags = []
    for c in cols:
        a_col = f"{c}_A" if f"{c}_A" in merged.columns else c
        b_col = f"{c}_B" if f"{c}_B" in merged.columns else c
        if a_col not in merged.columns or b_col not in merged.columns:
            continue
        a_vals = merged[a_col].astype("string").str.strip()
        b_vals = merged[b_col].astype("string").str.strip()
        is_diff = (a_vals != b_vals) & ~(a_vals.isna() & b_vals.isna())
        flag = f"DIFF__{c}"
        merged[flag] = is_diff
        diff_flags.append(flag)

    any_diff = merged[diff_flags].any(axis=1) if diff_flags else pd.Series([False]*len(merged))
    differences = merged[any_diff & merged["_key__"].notna()].copy()
    same = merged[(~any_diff) & merged["_key__"].notna()].copy()

    diff_view_cols = ["_key__"]
    for c in cols:
        a_col = f"{c}_A" if f"{c}_A" in merged.columns else c
        b_col = f"{c}_B" if f"{c}_B" in merged.columns else c
        d_col = f"DIFF__{c}"
        if a_col in differences.columns: diff_view_cols.append(a_col)
        if b_col in differences.columns: diff_view_cols.append(b_col)
        if d_col in differences.columns: diff_view_cols.append(d_col)

    differences_view = differences[diff_view_cols].copy().rename(columns={"_key__":"key"})

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        differences_view.to_excel(writer, index=False, sheet_name="Differences")
        same.drop(columns=["_key__"]).to_excel(writer, index=False, sheet_name="Same")
        not_found.drop(columns=["_key__"]).to_excel(writer, index=False, sheet_name="NotFound_in_B")
        if "_key__" in dup_report.columns:
            dup_report = dup_report.drop(columns=["_key__"])
        dup_report.to_excel(writer, index=False, sheet_name="Duplicates_in_B")

        meta = pd.DataFrame(
            {
                "item":[
                    "file_a","file_b","sheet_a","sheet_b",
                    "col_a_used","col_b_used","compare_cols",
                    "case_insensitive","blanks_dropped",
                    "count_differences","count_same","count_not_found","count_dup_rows_in_b"
                ],
                "value":[
                    file_a,file_b,str(sheet_a),str(sheet_b),
                    str(col_a),str(col_b),", ".join(map(str, cols)),
                    str(bool(case_insensitive)),str(not keep_blanks),
                    str(len(differences_view)),str(len(same)),str(len(not_found)),str(len(dup_report))
                ]
            }
        )
        meta.to_excel(writer, index=False, sheet_name="Meta")

    return {
        "out": out_path,
        "mode": "differences",
        "differences": len(differences_view),
        "same": len(same),
        "not_found": len(not_found),
        "dup_rows_in_b": len(dup_report),
        "compare_cols": cols,
    }
